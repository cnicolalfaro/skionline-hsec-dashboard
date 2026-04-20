from openpyxl import load_workbook
from pathlib import Path

path = Path(r"C:\Users\User\Downloads\TARJA ABRIL SK OBRA 250 SK  19-04-26 REV2.xlsx")
wb = load_workbook(path, data_only=True)
print("Hojas:", wb.sheetnames)

if 'TARJA' in wb.sheetnames:
    ws = wb['TARJA']
    rows = list(ws.iter_rows(values_only=True))
    # Buscar fila de headers
    header_row = None
    header_idx = 0
    for i, row in enumerate(rows[:15]):
        cells = [str(h).strip() if h is not None else '' for h in row]
        norm = [c.upper().replace('.', '').strip() for c in cells]
        if 'NOMBRE' in norm and ('RUT' in norm or 'A PATERNO' in norm):
            header_row = cells
            header_idx = i
            print(f"\nHeader en fila {i+1}")
            break

    if header_row:
        print("\n--- TODAS LAS COLUMNAS CON CONTENIDO ---")
        # Revisar columnas que tienen al menos un valor no vacío en datos
        data_rows = rows[header_idx+1:header_idx+21]
        for col_idx, col_name in enumerate(header_row):
            if not col_name:
                continue
            vals = [str(r[col_idx]).strip() for r in data_rows if len(r) > col_idx and r[col_idx] is not None and str(r[col_idx]).strip()]
            if vals:
                print(f"  [{col_idx:3d}] {repr(col_name):50s} -> ej: {repr(vals[0])}")

        print("\n--- PRIMERAS 3 FILAS DE DATOS ---")
        for i, row in enumerate(rows[header_idx+1:header_idx+4]):
            items = {header_row[j]: row[j] for j in range(min(len(header_row), len(row))) if header_row[j]}
            # Solo los no vacíos
            filled = {k: v for k, v in items.items() if v is not None and str(v).strip()}
            print(f"Fila {i+1}:", filled)
