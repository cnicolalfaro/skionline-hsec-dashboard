from __future__ import annotations

import json
import re
import unicodedata
from datetime import datetime
from difflib import SequenceMatcher
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = BASE_DIR / 'data'
OUTPUT_PATH = DATA_DIR / 'dashboard_data.js'
PEOPLE_INDEX_PATH = DATA_DIR / 'people_folders.json'
PREFERRED_EXCEL_NAME = 'REGISTRO_FINAL_CURSOS.xlsx'
SPECIAL_SHEETS = {'RESUMEN', 'TARJA', 'NO_LEGIBLE', 'DUPLICADOS'}

# URL base de SharePoint donde se replicarán las carpetas _PorPersona/<NOMBRE>.
# Las subcarpetas se construyen insertando "/<NOMBRE>" antes del "?".
PORPERSONA_SHAREPOINT_BASE = 'https://empresassk.sharepoint.com/:f:/r/sites/ICSK-HSEC/Documentos%20compartidos/05%20-%20Respaldo%20HSEC%20faenas/250%20-%20Mantenimiento%20M2%20y%20M3/Contrato%20250%20Dch/Sistema%20de%20Gesti%C3%B3n%20n%20contrato%204600030982/5-%20Respaldo%20documentaci%C3%B3n%20trabajadores/00_PORPERSONA?csf=1&web=1&e=d6nH9p'

# URL fija a la que apuntan las personas SIN carpeta propia (no tienen archivos en SharePoint),
# y también la sección "Sin match con TARJA" del dashboard.
PORPERSONA_SIN_MATCH_URL = 'https://empresassk.sharepoint.com/:f:/r/sites/ICSK-HSEC/Documentos%20compartidos/05%20-%20Respaldo%20HSEC%20faenas/250%20-%20Mantenimiento%20M2%20y%20M3/Contrato%20250%20Dch/Sistema%20de%20Gesti%C3%B3n%20n%20contrato%204600030982/5-%20Respaldo%20documentaci%C3%B3n%20trabajadores/00_PORPERSONA/TRABAJADOR%20SIN%20MATCH%20CON%20TARJA?csf=1&web=1&e=W1n3aM'
INDUCCION_COLS = [
    'CHARLA ADMINISTRATIVA',
    'CHARLA LEY KARIN',
    'CHARLA IRL',
    'CURSO DE ALTURA',
    'Inducción Persona Nueva',
    'Inducción a Geomecánica Básica',
    'Inducción Refugios Mineros',
    'Inducción Reglamento de Emergencias MCH',
    'Inducción al Reglamento de Tránsito en Mina',
    'Inducción de Cartillas de Evacuación',
    'Inducción General Mina Chuquicamata Subterránea',
    'Inducción uso Cintas de Confinamiento',
]


def format_date_str(value: str) -> str:
    """Convierte string de fecha openpyxl (YYYY-MM-DD HH:MM:SS) a dd/mm/yyyy si aplica."""
    try:
        dt = datetime.strptime(str(value).strip()[:10], '%Y-%m-%d')
        return dt.strftime('%d/%m/%Y')
    except (ValueError, TypeError):
        return str(value).strip()


def to_int(value: Any) -> int:
    try:
        return int(value)
    except (TypeError, ValueError):
        return 0


def find_excel_path() -> Path:
    search_paths = [
        DATA_DIR / PREFERRED_EXCEL_NAME,
        BASE_DIR / PREFERRED_EXCEL_NAME,
        BASE_DIR.parent / PREFERRED_EXCEL_NAME,
    ]

    preferred_existing = [path for path in search_paths if path.exists()]
    if preferred_existing:
        return max(preferred_existing, key=lambda item: item.stat().st_mtime)

    candidates: list[Path] = []
    for folder in [DATA_DIR, BASE_DIR, BASE_DIR.parent]:
        candidates.extend(folder.glob('*.xlsx'))

    candidates = sorted(set(candidates), key=lambda item: item.stat().st_mtime, reverse=True)
    if candidates:
        return candidates[0]

    raise FileNotFoundError('No se encontró ningún archivo Excel para procesar.')


def normalize_text(value: Any) -> str:
    text = '' if value is None else str(value)
    text = unicodedata.normalize('NFD', text)
    text = ''.join(char for char in text if unicodedata.category(char) != 'Mn')
    text = re.sub(r'\bfalta\s*firma\b', ' ', text, flags=re.IGNORECASE)
    text = re.sub(r'\bfaltafirma\b', ' ', text, flags=re.IGNORECASE)
    text = re.sub(r'\bforms\b', ' ', text, flags=re.IGNORECASE)
    text = re.sub(r'\birl\s*general\b', ' ', text, flags=re.IGNORECASE)
    text = re.sub(r'\birl\s*especifica\b', ' ', text, flags=re.IGNORECASE)
    text = re.sub(r'[^A-Za-z ]+', ' ', text).lower()
    return ' '.join(text.split())


def split_name_tokens(value: Any) -> list[str]:
    return [token for token in normalize_text(value).split() if len(token) > 1]


def fingerprint(*values: Any) -> str:
    words: list[str] = []
    for value in values:
        words.extend(split_name_tokens(value))
    return ' '.join(sorted(words))


def slug_token(value: Any) -> str:
    """Convierte un token a A-Z/0-9 (sin acentos, mayúsculas)."""
    if value is None:
        return ''
    text = unicodedata.normalize('NFD', str(value))
    text = ''.join(c for c in text if unicodedata.category(c) != 'Mn')
    return re.sub(r'[^A-Z0-9]', '', text.upper())


def build_person_folder(nombre: Any, paterno: Any, materno: Any, rut: Any) -> str:
    """PATERNO_MATERNO_NOMBRE1_NOMBRE2_RUT (omite partes vacías)."""
    parts: list[str] = []
    pat = slug_token(paterno)
    if pat:
        parts.append(pat)
    mat = slug_token(materno)
    if mat:
        parts.append(mat)
    nombres = [n for n in str(nombre or '').split() if n.strip()]
    for n in nombres[:2]:
        slug = slug_token(n)
        if slug:
            parts.append(slug)
    rut_clean = re.sub(r'[^0-9Kk]', '', str(rut or '')).upper()
    if rut_clean:
        parts.append(rut_clean)
    return '_'.join(parts)


def build_person_folder_url(folder_name: str) -> str:
    """Inserta el segmento /<folder_name> en la URL base de SharePoint."""
    if not folder_name or not PORPERSONA_SHAREPOINT_BASE:
        return ''
    base = PORPERSONA_SHAREPOINT_BASE
    # Insertar antes del '?' (si existe)
    if '?' in base:
        path_part, qs = base.split('?', 1)
        return f'{path_part}/{folder_name}?{qs}'
    return f'{base}/{folder_name}'


def load_people_folders_index() -> dict[str, str]:
    """Lee SKIONLINE/data/people_folders.json y devuelve {rutCanon: folder}."""
    if not PEOPLE_INDEX_PATH.exists():
        return {}
    try:
        with open(PEOPLE_INDEX_PATH, 'r', encoding='utf-8') as f:
            data = json.load(f)
    except (json.JSONDecodeError, OSError):
        return {}
    index: dict[str, str] = {}
    for entry in data.values() if isinstance(data, dict) else []:
        if not isinstance(entry, dict):
            continue
        canon = str(entry.get('rutCanon') or '').strip()
        folder = str(entry.get('folder') or '').strip()
        if canon and folder:
            index[canon] = folder
    return index


def names_match(evidence_name: Any, nombre: Any, paterno: Any, materno: Any) -> bool:
    evidence_tokens = split_name_tokens(evidence_name)
    nombre_tokens = split_name_tokens(nombre)
    paterno_tokens = split_name_tokens(paterno)
    materno_tokens = split_name_tokens(materno)

    if not evidence_tokens:
        return False

    full_tokens = nombre_tokens + paterno_tokens + materno_tokens
    if evidence_tokens and set(evidence_tokens) == set(full_tokens):
        return True

    surname_tokens = [*paterno_tokens[:1], *materno_tokens[:1]]
    if not surname_tokens:
        return False

    surname_match_count = 0
    for surname in surname_tokens:
        if any(
            token == surname or SequenceMatcher(None, token, surname).ratio() >= 0.82
            for token in evidence_tokens
        ):
            surname_match_count += 1

    if surname_match_count < len(surname_tokens):
        return False

    if any(token in evidence_tokens for token in nombre_tokens):
        return True

    if nombre_tokens:
        evidence_first = evidence_tokens[0]
        similarity = max((SequenceMatcher(None, evidence_first, token).ratio() for token in nombre_tokens), default=0)
        return similarity >= 0.80

    return False


def collect_person_evidence(evidence_entries: list[dict[str, Any]], nombre: Any, paterno: Any, materno: Any, rut: Any = '') -> dict[str, set[str]]:
    courses: set[str] = set()
    flags: set[str] = set()
    notes: set[str] = set()
    rut_matches: set[str] = set()
    target_rut = clean_rut(rut)

    for entry in evidence_entries:
        entry_ruts = {clean_rut(value) for value in entry.get('rutCandidates', []) if clean_rut(value)}
        matches_by_rut = bool(target_rut and target_rut in entry_ruts)

        if not matches_by_rut and not names_match(entry.get('name'), nombre, paterno, materno):
            continue

        if matches_by_rut:
            rut_matches.add(target_rut)
        if entry.get('course'):
            courses.add(entry['course'])
        if entry.get('flag'):
            flags.add(entry['flag'])
        if entry.get('note'):
            notes.add(entry['note'])

    return {'courses': courses, 'flags': flags, 'notes': notes, 'rutMatches': rut_matches}


def get_cell_value(row: tuple[Any, ...], headers: list[str], column_name: str, default: str = '') -> str:
    if column_name not in headers:
        return default
    idx = headers.index(column_name)
    if len(row) <= idx or row[idx] is None:
        return default
    return str(row[idx]).strip()


def is_evidence_sheet(sheet_name: str) -> bool:
    return sheet_name not in SPECIAL_SHEETS


# Hojas que no se muestran en el dashboard (totalmente ocultas de gráficos y tabla resumen)
# Solo IRL_GENERAL_FORMS queda oculto; los cursos de mutual ahora se muestran como columnas.
HIDDEN_FROM_DASHBOARD = {'IRL_GENERAL_FORMS'}

# Nombres completos amigables para la visualización en el dashboard
COURSE_FULL_NAMES = {
    'AYB': 'Aislación y Bloqueo',
    'EPP': 'Uso y Mantención de EPP',
    'EXT': 'Manejo de Extintores Portátiles',
    'OPR': 'Orientación a la Prevención de Riesgos',
    'PA': 'Primeros Auxilios',
    'IRL_GENERAL': 'IRL General',
    'IRL GENERAL': 'IRL General',
    'IRL_ESPECIFICA': 'IRL Específica',
    'IRL ESPECIFICA': 'IRL Específica',
    'EVALUACIONES_IRL': 'Evaluaciones IRL',
    'EVALUACIONES IRL': 'Evaluaciones IRL',
    'CAD': 'Conducción a la Defensiva',
}


def format_course_name(sheet_name: str) -> str:
    return sheet_name.replace('_', ' ').strip()


def display_course_name(sheet_name: str) -> str:
    """Devuelve el nombre completo/amigable del curso para el dashboard."""
    key = sheet_name.strip()
    if key in COURSE_FULL_NAMES:
        return COURSE_FULL_NAMES[key]
    formatted = format_course_name(key)
    return COURSE_FULL_NAMES.get(formatted, formatted)


def format_rut(value: Any) -> str:
    text = '' if value is None else str(value).strip()
    cleaned = re.sub(r'[^0-9kK]', '', text)
    if len(cleaned) < 2:
        return text
    return f'{cleaned[:-1]}-{cleaned[-1].upper()}'


def clean_rut(value: Any) -> str:
    cleaned = re.sub(r'[^0-9kK]', '', '' if value is None else str(value)).upper()
    if len(cleaned) >= 2:
        body = cleaned[:-1].lstrip('0') or '0'
        return body + cleaned[-1]
    return cleaned


def extract_rut_candidates(value: Any) -> list[str]:
    text = '' if value is None else str(value)
    matches = re.findall(r'(?<!\d)(\d{7,8}[\-–]?[0-9kK])(?!\d)', text)
    candidates: list[str] = []
    for match in matches:
        cleaned = clean_rut(match)
        if len(cleaned) >= 8:
            candidates.append(cleaned)
    return list(dict.fromkeys(candidates))


def get_first_available_cell(row: tuple[Any, ...], headers: list[str], column_names: list[str], default: str = '') -> str:
    normalized_headers = [header.upper().replace('.', '').strip() for header in headers]
    for column_name in column_names:
        normalized_name = column_name.upper().replace('.', '').strip()
        if normalized_name not in normalized_headers:
            continue
        idx = normalized_headers.index(normalized_name)
        if len(row) <= idx or row[idx] is None:
            continue
        value = str(row[idx]).strip()
        if value:
            return value
    return default


def extract_headers_and_rows(raw_rows: list[tuple[Any, ...]]) -> tuple[list[str], list[tuple[Any, ...]]]:
    if not raw_rows:
        return [], []

    for idx, raw_header in enumerate(raw_rows[:15]):
        headers = [str(h).strip() if h is not None else '' for h in raw_header]
        normalized = [header.upper().replace('.', '').strip() for header in headers]
        # Aceptar diversas variantes del header de apellido paterno: "A PATERNO",
        # "APELLIDO PATERNO", "APELLIDOS" (TARJA 27-04-26), o cualquier columna
        # que contenga PATERNO/APELLIDO. Robusto a futuros cambios de la TARJA.
        has_apellido = (
            'A PATERNO' in normalized
            or 'APELLIDOS' in normalized
            or any('PATERNO' in h or h.startswith('APELLIDO') for h in normalized)
        )
        if 'NOMBRE' in normalized and ('RUT' in normalized or has_apellido):
            return headers, raw_rows[idx + 1:]

    headers = [str(h).strip() if h is not None else '' for h in raw_rows[0]]
    return headers, raw_rows[1:]


def _safe_iter_rows(ws, max_empty_run: int = 30, hard_limit: int = 5000) -> list[tuple[Any, ...]]:
    """Lee filas de una hoja deteniéndose al encontrar una racha de filas vacías.

    Esto evita quedarse colgado cuando openpyxl (read_only) reporta dimensiones
    infladas por pivot caches u otros artefactos del archivo.
    """
    rows: list[tuple[Any, ...]] = []
    empty_run = 0
    for row in ws.iter_rows(values_only=True):
        if row is None:
            empty_run += 1
        elif all(v is None or (isinstance(v, str) and not v.strip()) for v in row):
            empty_run += 1
        else:
            empty_run = 0
        rows.append(row)
        if empty_run >= max_empty_run or len(rows) >= hard_limit:
            break
    return rows


def find_external_tarja_workbook():
    search_folders = [BASE_DIR, BASE_DIR.parent, DATA_DIR]

    candidates: list[Path] = []
    for folder in search_folders:
        candidates.extend(folder.glob('TARJA*.xlsx'))

    # Ordenar por fecha de modificación descendente (más reciente primero)
    candidates = sorted(set(candidates), key=lambda p: p.stat().st_mtime, reverse=True)

    for path in candidates:
        try:
            wb = load_workbook(path, data_only=True, read_only=True)
        except Exception:
            continue
        if 'TARJA' in wb.sheetnames:
            print(f'Usando TARJA externo: {path.name}')
            return wb

    return None


def load_external_rut_lookup() -> dict[str, str]:
    lookup: dict[str, str] = {}
    wb = find_external_tarja_workbook()
    if wb is None:
        return lookup

    ws = wb['TARJA']
    raw_rows = _safe_iter_rows(ws)
    headers, rows = extract_headers_and_rows(raw_rows)
    if not headers or not rows:
        return lookup

    if 'RUT' not in [header.upper().replace('.', '').strip() for header in headers]:
        return lookup

    for row in rows:
        if not row or all(value is None or str(value).strip() == '' for value in row):
            continue

        nombre = get_first_available_cell(row, headers, ['NOMBRE'])
        paterno = get_first_available_cell(row, headers, ['A. PATERNO', 'APELLIDO PATERNO', 'APELLIDOS', 'APELLIDO'])
        materno = get_first_available_cell(row, headers, ['A. MATERNO', 'APELLIDO MATERNO'])
        rut = format_rut(get_first_available_cell(row, headers, ['RUT']))
        person_key = fingerprint(nombre, paterno, materno)

        if person_key and rut:
            lookup[person_key] = rut

    return lookup


def build_tarja_records(wb, evidence_entries: list[dict[str, Any]], external_rut_lookup: dict[str, str]) -> tuple[list[dict[str, Any]], int, list[str]]:
    # Siempre preferir el TARJA externo (archivo TARJA*.xlsx en carpeta superior),
    # porque es la fuente autoritativa y más actualizada de los turnos diarios.
    # Solo caer al TARJA embebido del REGISTRO si no hay externo disponible.
    external_wb = find_external_tarja_workbook()
    if external_wb is not None and 'TARJA' in external_wb.sheetnames:
        ws = external_wb['TARJA']
    elif 'TARJA' in wb.sheetnames:
        ws = wb['TARJA']
    else:
        return [], 0, []
    raw_rows = _safe_iter_rows(ws)
    headers, rows = extract_headers_and_rows(raw_rows)
    if not headers or not rows:
        return [], 0, []

    # Índice de carpetas SharePoint generadas por generar_porpersona.py.
    # Si una persona no aparece aquí, es porque no tiene archivos asociados:
    # el botón 📁 apuntará a la carpeta global "TRABAJADOR SIN MATCH CON TARJA".
    folders_index = load_people_folders_index()

    # Detectar columnas con fechas de turno (datetime o strings comunes)
    shift_cols: list[int] = []
    shift_dates_iso: list[str] = []
    dmy_pattern = re.compile(r'^(\d{2})-(\d{2})-(\d{4})$')
    ymd_pattern = re.compile(r'^(\d{4})-(\d{2})-(\d{2})')
    for idx, h in enumerate(headers):
        iso = ''
        if isinstance(h, datetime):
            iso = h.strftime('%Y-%m-%d')
        else:
            s = str(h).strip()
            m = dmy_pattern.match(s)
            if m:
                dd, mm, yyyy = m.groups()
                iso = f'{yyyy}-{mm}-{dd}'
            else:
                m2 = ymd_pattern.match(s)
                if m2:
                    iso = f'{m2.group(1)}-{m2.group(2)}-{m2.group(3)}'
        if iso:
            shift_cols.append(idx)
            shift_dates_iso.append(iso)

    records: list[dict[str, Any]] = []
    sin_registros = 0

    for row in rows:
        if not row or all(value is None or str(value).strip() == '' for value in row):
            continue

        nombre = get_cell_value(row, headers, 'NOMBRE')
        paterno = get_first_available_cell(row, headers, ['A. PATERNO', 'APELLIDO PATERNO', 'APELLIDOS', 'APELLIDO'])
        materno = get_first_available_cell(row, headers, ['A. MATERNO', 'APELLIDO MATERNO'])
        especialidad = get_cell_value(row, headers, 'ESPECIALIDAD', 'Sin especialidad informada')
        estado_tarja = get_cell_value(row, headers, 'ESTADO', 'Sin estado')

        categoria = get_first_available_cell(row, headers, ['CATEGORIA'])
        turno = get_first_available_cell(row, headers, ['TURNO'])
        acr_sucal = get_first_available_cell(row, headers, ['ACR. SUCAL', 'ACR SUCAL', 'ACRSUCAL']).strip()
        correo = get_first_available_cell(row, headers, ['CORREO', 'EMAIL', 'E-MAIL', 'MAIL']).strip()
        fono = get_first_available_cell(row, headers, ['FONO', 'TELEFONO', 'TELÉFONO', 'CELULAR']).strip()
        cert_final_raw = get_first_available_cell(row, headers, ['CERTIFICADO FINAL'])
        cert_final = format_date_str(cert_final_raw) if cert_final_raw else ''
        if not cert_final:
            pendiente_raw = get_first_available_cell(row, headers, ['PENDIENTE'])
            if pendiente_raw and 'PENDIENTE' in pendiente_raw.upper():
                cert_final = 'PENDIENTE'
        examen_salud_raw = get_first_available_cell(row, headers, ['ESTATUS VENCIMIENTO EXAMEN DE SALUD'])
        examen_salud = format_date_str(examen_salud_raw) if examen_salud_raw else ''

        inducciones_ok: list[str] = []
        for ind_col in INDUCCION_COLS:
            ind_val = get_first_available_cell(row, headers, [ind_col])
            if ind_val and ind_val.upper().strip() not in ('PENDIENTE', '-', 'NO', ''):
                inducciones_ok.append(ind_col)

        nombre_mostrado = ' '.join(part for part in [paterno, materno, nombre] if part).strip() or 'Sin nombre'
        person_key = fingerprint(nombre, paterno, materno)
        rut_tarja = format_rut(get_first_available_cell(row, headers, ['RUT', 'R.U.T', 'R U T'], ''))
        rut = rut_tarja or external_rut_lookup.get(person_key, '')

        # Resolver carpeta SharePoint: si la persona tiene archivos copiados a
        # _PorPersona/, usar su carpeta exacta; si no, apuntar a la carpeta
        # global "TRABAJADOR SIN MATCH CON TARJA".
        rut_canon = clean_rut(rut)
        existing_folder = folders_index.get(rut_canon, '') if rut_canon else ''
        if existing_folder:
            folder_name = existing_folder
            folder_url = build_person_folder_url(existing_folder)
        else:
            folder_name = ''
            folder_url = PORPERSONA_SIN_MATCH_URL

        evidence = collect_person_evidence(evidence_entries, nombre, paterno, materno, rut)
        courses = sorted(evidence.get('courses', set()))
        flags = evidence.get('flags', set())
        notes = evidence.get('notes', set())

        tarja_extra_parts: list[str] = []
        if acr_sucal:
            tarja_extra_parts.append(f'ACR: {acr_sucal}')
        if categoria:
            tarja_extra_parts.append(f'Cat: {categoria}')
        if turno:
            tarja_extra_parts.append(f'Turno: {turno}')
        if examen_salud:
            tarja_extra_parts.append(f'Exam. salud: {examen_salud}')
        tarja_extra = (' · ' + ' · '.join(tarja_extra_parts)) if tarja_extra_parts else ''

        if not courses and not flags:
            estado = 'No hay registros cargados, confirmar en portales'
            status_key = 'sin-registros'
            detalle = f'{especialidad} · Estado TARJA: {estado_tarja}{tarja_extra}'
            sin_registros += 1
        else:
            if courses:
                estado = 'Con registros'
                status_key = 'con-registros'
            elif 'no-legible' in flags:
                estado = 'No legible'
                status_key = 'no-legible'
            else:
                estado = 'Duplicado'
                status_key = 'duplicado'

            observaciones = []
            if 'duplicado' in flags:
                observaciones.append('incluye duplicados')
            if 'no-legible' in flags:
                observaciones.append('incluye no legibles')
            if 'irl-forms' in notes:
                observaciones.append('respaldo vía Forms; se recomienda validar el archivo original en la carpeta documental')
            extra = f" · {', '.join(observaciones)}" if observaciones else ''
            detalle = f'{especialidad} · Estado TARJA: {estado_tarja}{extra}{tarja_extra}'

        records.append({
            'nombre': nombre_mostrado,
            'rut': rut,
            'folderName': folder_name,
            'folderUrl': folder_url,
            'cursos': ', '.join(courses) if courses else '-',
            'courseList': courses,
            'estado': estado,
            'statusKey': status_key,
            'detalle': detalle,
            'certFinal': cert_final,
            'examenSalud': examen_salud,
            'categoria': categoria,
            'turno': turno,
            'acrSucal': acr_sucal,
            'correo': correo,
            'fono': fono,
            'inducionesOk': len(inducciones_ok),
            'inducionesTotal': len(INDUCCION_COLS),
            'inducionesList': inducciones_ok,
            'shifts': [
                (str(row[c]).strip().upper() if c < len(row) and row[c] is not None else '')
                for c in shift_cols
            ],
        })

    records.sort(key=lambda item: normalize_text(item['nombre']))
    return records, sin_registros, shift_dates_iso


SIN_MATCH_DIRS = {
    'IRL GENERAL': 'IRL General',
    'IRL_GENERAL_FORMS': 'IRL General (Forms)',
    'IRL ESPECIFICA': 'IRL Específica',
    'EVALUACIONES IRL': 'Evaluaciones IRL',
}


def collect_sin_match_files(tarja_rut_set: set[str] | None = None) -> list[dict[str, Any]]:
    """Escanea las carpetas IRL en disco buscando archivos sin match con la TARJA.

    Incluye:
      1. Archivos con prefijo `_SIN_MATCH_` en su nombre (marcados explícitamente
         por el flujo de renombrado/duplicados).
      2. Archivos cuyo RUT extraído del nombre NO existe en la TARJA actual
         (la persona tiene documentos cargados pero no aparece en la planilla).
    """
    diplomas_root = BASE_DIR.parent / 'renombrar' / 'Diplomas'
    if not diplomas_root.exists():
        return []

    rut_pattern = re.compile(r'(?<!\d)(\d{7,9}[kK]?)(?!\d)')
    tarja_rut_set = tarja_rut_set or set()
    results: list[dict[str, Any]] = []
    seen: set[tuple[str, str]] = set()

    for folder_name, label in SIN_MATCH_DIRS.items():
        folder = diplomas_root / folder_name
        if not folder.is_dir():
            continue
        for entry in sorted(folder.iterdir(), key=lambda p: p.name.lower()):
            if not entry.is_file():
                continue
            name = entry.name
            base = entry.stem
            upper = name.upper()
            is_explicit = '_SIN_MATCH_' in upper

            if is_explicit:
                match = re.split(r'_SIN_MATCH_', base, flags=re.IGNORECASE)
                leftover = match[-1] if match else base
                ruts_found = rut_pattern.findall(leftover)
                rut_raw = ruts_found[0] if ruts_found else ''
                hint_text = re.sub(r'[_\-]+', ' ', leftover).strip()
                if rut_raw:
                    hint_text = hint_text.replace(rut_raw, '').strip(' _-')
                hint = hint_text or 'Sin nombre detectable'
            else:
                # Buscar RUT en todo el nombre del archivo
                ruts_found = rut_pattern.findall(base)
                if not ruts_found:
                    continue
                rut_raw = ruts_found[0]
                rut_canon = clean_rut(rut_raw)
                # Solo incluir si el RUT NO está en la TARJA
                if not rut_canon or rut_canon in tarja_rut_set:
                    continue
                # Limpiar nombre del archivo para usar como hint
                hint_text = re.sub(r'[_\-]+', ' ', base).strip()
                hint_text = re.sub(r'\b' + re.escape(rut_raw) + r'\b', '', hint_text).strip()
                hint = hint_text or 'RUT no presente en TARJA'

            rut_formatted = format_rut(rut_raw) if rut_raw else ''
            key = (folder_name, name)
            if key in seen:
                continue
            seen.add(key)

            results.append({
                'folder': folder_name,
                'folderLabel': label,
                'file': name,
                'rut': rut_formatted,
                'hint': hint,
            })
    return results


def main() -> None:
    excel_path = find_excel_path()
    wb = load_workbook(excel_path, data_only=True)

    summary_rows: list[dict[str, Any]] = []
    summary_map: dict[str, dict[str, Any]] = {}

    if 'RESUMEN' in wb.sheetnames:
        ws = wb['RESUMEN']
        rows = list(ws.iter_rows(values_only=True))
        headers = [str(h).strip() if h is not None else '' for h in rows[0]]

        for row in rows[1:]:
            if not any(cell is not None for cell in row):
                continue
            item = {headers[i]: row[i] for i in range(min(len(headers), len(row)))}
            curso = str(item.get('CURSO', '')).strip()
            # Ocultar cursos no deseados del dashboard (se mantienen en el Excel)
            if curso in HIDDEN_FROM_DASHBOARD:
                continue
            total = to_int(item.get('TOTAL_ARCHIVOS'))
            unicos_raw = item.get('UNICOS')
            unicos = None if unicos_raw in (None, '') else to_int(unicos_raw)
            parsed = {'curso': display_course_name(curso), 'total': total, 'unicos': unicos}
            summary_rows.append(parsed)
            summary_map[curso] = parsed

    course_totals: list[dict[str, Any]] = []
    evidence_entries: list[dict[str, Any]] = []

    for sheet_name in wb.sheetnames:
        if sheet_name in {'RESUMEN', 'TARJA'}:
            continue        # Ocultar hojas no deseadas del dashboard (siguen existiendo en el Excel)
        if sheet_name in HIDDEN_FROM_DASHBOARD:
            continue
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        headers = [str(h).strip() if h is not None else '' for h in rows[0]]
        data_rows = [
            row for row in rows[1:]
            if any(cell is not None and str(cell).strip() != '' for cell in row)
        ]

        if is_evidence_sheet(sheet_name):
            course_totals.append({
                'curso': display_course_name(sheet_name),
                'total': summary_map.get(sheet_name, {}).get('total', len(data_rows))
            })

        if 'NOMBRE_PERSONA' not in headers:
            continue

        for row in data_rows:
            nombre_persona = get_cell_value(row, headers, 'NOMBRE_PERSONA')
            nombre_archivo = get_cell_value(row, headers, 'NOMBRE_ARCHIVO')
            person_key = fingerprint(nombre_persona)
            rut_candidates = extract_rut_candidates(nombre_persona) + extract_rut_candidates(nombre_archivo)
            # Si no hay tokens nombre (p.ej. "IRL GENERAL 17875209K") pero sí hay RUT,
            # igualmente agregar la entry para que cruce por RUT con la TARJA.
            if not person_key and not rut_candidates:
                continue

            course_name = format_course_name(sheet_name)
            entry = {
                'name': nombre_persona,
                'rutCandidates': rut_candidates,
                'course': '',
                'flag': '',
                'note': '',
            }

            if is_evidence_sheet(sheet_name):
                entry['course'] = course_name
                if course_name == 'IRL GENERAL FORMS':
                    entry['note'] = 'irl-forms'
            elif sheet_name == 'NO_LEGIBLE':
                entry['flag'] = 'no-legible'
            elif sheet_name == 'DUPLICADOS':
                entry['flag'] = 'duplicado'

            evidence_entries.append(entry)

    external_rut_lookup = load_external_rut_lookup()
    records, sin_registros, shift_dates = build_tarja_records(wb, evidence_entries, external_rut_lookup)
    tarja_rut_set = {clean_rut(r.get('rut', '')) for r in records if r.get('rut')}
    tarja_rut_set.discard('')
    sin_match_files = collect_sin_match_files(tarja_rut_set)

    total_archivos = summary_map.get('TOTAL', {}).get('total', sum(item['total'] for item in course_totals))
    documentos_unicos = sum(item['unicos'] for item in summary_rows if isinstance(item.get('unicos'), int))
    duplicados = summary_map.get('DUPLICADOS', {}).get('total', 0)
    no_legibles = summary_map.get('NO_LEGIBLE', {}).get('total', 0)
    trabajadores_tarja = len(records)
    con_registros = max(trabajadores_tarja - sin_registros, 0)
    aprobados_tarja = sum(1 for r in records if 'APROBADO' in (r.get('certFinal') or '').upper())
    pendientes_tarja = sum(1 for r in records if 'PENDIENTE' in (r.get('certFinal') or '').upper())

    curso_top = max(course_totals, key=lambda item: item['total']) if course_totals else {'curso': '-', 'total': 0}
    payload = {
        'generatedAt': datetime.now().strftime('%d/%m/%Y %H:%M'),
        'sourceFile': excel_path.name,
        'kpis': {
            'totalArchivos': total_archivos,
            'documentosUnicos': documentos_unicos,
            'duplicados': duplicados,
            'noLegibles': no_legibles,
            'trabajadoresTarja': trabajadores_tarja,
            'conRegistros': con_registros,
            'sinRegistros': sin_registros,
        },
        'courseTotals': course_totals,
        'statusBreakdown': [
            {'label': 'Con registros', 'value': con_registros, 'color': '#51b847'},
            {'label': 'Sin registros', 'value': sin_registros, 'color': '#ff7a59'},
            {'label': 'No legibles', 'value': no_legibles, 'color': '#f53b4d'},
        ],
        'summaryRows': summary_rows,
        'records': records,
        'shiftDates': shift_dates,
        'sinMatchFiles': sin_match_files,
        'sinMatchFolderUrl': PORPERSONA_SIN_MATCH_URL,
        'irlFormsNote': '',
        'accessLinks': [
            {'label': 'Evidencias de certificaciones', 'url': 'https://empresassk.sharepoint.com/:f:/r/sites/ICSK-HSEC/Documentos%20compartidos/05%20-%20Respaldo%20HSEC%20faenas/250%20-%20Mantenimiento%20M2%20y%20M3/Contrato%20250%20Dch/Sistema%20de%20Gesti%C3%B3n%20n%20contrato%204600030982/3-%20Registros%20capacitaciones%20-%20difusiones/3-%20Evidencias%20de%20Certificaciones?csf=1&web=1&e=ovhAFT'},
            {'label': 'Procedimientos contrato 982', 'url': 'https://empresassk.sharepoint.com/:f:/r/sites/ICSK-HSEC/Documentos%20compartidos/05%20-%20Respaldo%20HSEC%20faenas/250%20-%20Mantenimiento%20M2%20y%20M3/Contrato%20250%20Dch/Sistema%20de%20Gesti%C3%B3n%20n%20contrato%204600030982/2-%20Procedimientos?csf=1&web=1&e=IsTxdJ'},
            {'label': 'Procedimientos contrato 984', 'url': 'https://empresassk.sharepoint.com/:f:/r/sites/ICSK-HSEC/Documentos%20compartidos/05%20-%20Respaldo%20HSEC%20faenas/250%20-%20Mantenimiento%20M2%20y%20M3/Contrato%20250%20Dch/Sistema%20de%20Gesti%C3%B3n%20n%20contrato%204600030984/2-%20Procedimientos?csf=1&web=1&e=RJlP81'},
            {'label': 'Check list actualizados', 'url': 'https://empresassk.sharepoint.com/:f:/r/sites/ICSK-HSEC/Documentos%20compartidos/05%20-%20Respaldo%20HSEC%20faenas/250%20-%20Mantenimiento%20M2%20y%20M3/Contrato%20250%20Dch/Sistema%20de%20Gesti%C3%B3n%20n%20contrato%204600030982/4-%20Formatos%20terreno/01-%20Check%20list/Check%20List%20Actualizados?csf=1&web=1&e=5arIgO'}
        ],
        'insights': [
            {'title': 'Curso con mayor volumen', 'detail': f"{curso_top['curso']} concentra {curso_top['total']} archivos."},
            {'title': 'Cruce con TARJA', 'detail': f'Se detectaron {sin_registros} trabajadores sin evidencias cargadas en el sistema.'},
            {'title': 'Calidad del registro', 'detail': f'Se identificaron {duplicados} duplicados y {no_legibles} documentos no legibles.'},
            {'title': 'Certificados finales (TARJA)', 'detail': f'{aprobados_tarja} trabajadores aprobados y {pendientes_tarja} pendientes según hoja TARJA.'},
        ],
    }

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT_PATH.write_text(
        'window.DASHBOARD_DATA = ' + json.dumps(payload, ensure_ascii=False, indent=2) + ';',
        encoding='utf-8',
    )
    print(f'Dashboard data generated at: {OUTPUT_PATH} using {excel_path.name}')


if __name__ == '__main__':
    main()
